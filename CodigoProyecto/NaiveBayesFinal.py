import re
import ssl
import openpyxl
from nltk import WordNetLemmatizer
from nltk.corpus import stopwords
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.naive_bayes import MultinomialNB  # Importar Naive Bayes
import joblib
import nltk
import matplotlib.pyplot as plt
import numpy as np
import re
import csv
from collections import Counter
from unidecode import unidecode

# Deshabilitar la verificación de certificados SSL (solo fue utilizada para instalar ciertas funciones de la libreria nltk)
ssl._create_default_https_context = ssl._create_unverified_context

# Cargar el diccionario principal
palabras_muy_bueno = ["primera", "futura", "bien", "amo", "correcto", "excelente", "fe",
                      "ayuda", "corazones", "corazon", "intelectual", "mejor", "excelentes", "#esclaudia️", "grandeza",
                      "gano", "bendiga", "monitorear", "bendiciones", "apoyo", "Te amo", "salvar", "felicidades", "vamos",
                      "ganara", "gana", "votare", "respetos", "respetar", "demostrando", "recta", "culta", "salud", "vamos",
                      "maestria", "doctora", "estudiada", "fortaleza", "preparada", "gran", "respetuoso", "coherente",
                      "excepcional", "alegria", "esperanza", "amor", "union", "fuerza", "dama", "elegancia", "indudablemente",
                      "informada", "preparada", "feminista", "proxima", "diferencia", "inteligencia", "sublime", "clara",
                      "detallada", "mujeron", "primera", "preparada", "esperanzador", "inspirador", "comprometido",
                      "perspicaz", "transformador", "visionario", "innovador", "empoderado", "solidario", "persuasivo",
                      "revolucionario", "brillante", "perspicuo", "admirable", "aclamado", "autentico", "carismatico",
                      "prometedor", "inspiracional", "impresionante"]

palabra_muy_mal = ["mala persona", "mala", "pena", "asquerosa", "falta", "farza", "estúpidos", "estúpida", "urge",
                   "sangre", "ridículo", "ridícula", "robara", "corrupcion", "corrupta", "narco", "fea", "amargura",
                   "asquerosa", "corrupta", "enfermita", "enfrema", "botarga", "falsa", "vieja", "mentira", "miedo",
                   "pena", "eliminar", "quitar", "ridicula", "neta", "desesperacion", "cuesta", "secta", "ja", "jaja",
                   "jajaja", "jajajajaj", "broma", "sangre", "corriente", "naca", "nacos", "perder", "perdedora",
                   "parodia", "ridiculos", "piche", "gorda", "deshonesto", "desilusionante", "desastroso", "manipulador",
                   "desconectado", "despectivo", "contraproducente", "desalentador", "incompetente", "desafiante",
                   "desgarrador", "decepcionante", "desconcertante", "despiadado", "irrespetuoso", "irresponsable",
                   "prejuicioso"]

diccionario_principal = {"palabras_muy_bueno": palabras_muy_bueno, "palabra_muy_mal": palabra_muy_mal}


stop_words = {
    'la', 'el', 'y', 'yo', 'ella', 'es', 'a', 'de', 'que', 'en', 'un', 'por', 'con', 'no', 'una', 'su', 'para', 'es',
    'como', 'esta', 'pero', 'tiene', 'solo', 'porque', 'vamos', 'duda', 'apoyo', 'mucho', 'siempre', 'sera', 'cuando',
    'nada', 'hacer', 'puede', 'tambien', 'quien', 'quiere', 'sabe', 'nunca', 'entre', 'ahora', 'estamos', 'total',
    'hace', 'tener', 'desde', 'hasta', 'estoy', 'algo', 'sobre', 'igual', 'estan', 'buenas', 'pues', 'esto',
    'forma', 'nota', 'estos', 'hecho', 'ojala', 'demuestra', 'claro', 'poder',
    'super', 'respuesta', 'tiempo', 'donde', 'seguro', 'favor', 'sigue', 'cabe', 'quedo', 'aqui', 'tienes',
    'cuenta', 'queda', 'tipo', 'muchas', 'otro', 'otros', 'correcto', 'gusto', 'tengo', 'mexico', 'presidente',
    'mujer', 'bien', 'lopez', 'este', 'todo', 'mejor', 'amlo', 'gran',
    'sheinbaum', 'pais', 'nuestro', 'todos', 'joaquin', 'candidata',
    'eres', 'usted', 'respuestas', 'xochitl', 'mexicanos', 'pueblo', 'preguntas', 'senora', 'senor',
    'otra', 'morena', 'buena', 'obrador', 'anos', 'persona', 'chayote', 'gobierno',
    'diferencia', 'dios', 'mucha', 'respetos', 'periodista', 'habla', 'arriba', 'gente', 'chayoteros',
    'saludos', 'continuidad', 'orgullo', 'mismo', 'gracias', 'verdad', 'toda', 'estar', 'grande',
    'presidencia', 'excelentes', 'plan', 'preparacion', 'hablar', 'jueces', 'contra',
    'politicos', 'proyecto', 'tienen', 'seguir', 'politica', 'bravo', 'super',
    'respuesta', 'tiempo', 'donde', 'seguro', 'favor', 'sigue', 'cabe', 'galvez',
    'unidos', 'sheimbaun', 'buen', 'piso', 'mundo', 'manera', 'patria', 'quedo', 'aqui', 'sheimbaum', 'falta',
    'contestar', 'dice', 'tienes', 'cuenta', 'corrupcion', 'gusta', 'opcion', 'dejo', 'muchas',
    'ciro', 'queda', 'tipo', 'entrevistas', 'teacher', 'temas', 'otro', 'trata',
    'anteriores', 'educada', 'risa', 'pudo', 'tenemos', 'otros', 'correcto', 'estados', 'oposicion', 'nuestros',
    'futuro', 'nivel', 'importante', 'gusto', 'tengo'
}

# Función para preprocesar el texto
def preprocess_text(text):
    # Eliminar enlaces, menciones y caracteres no alfabéticos
    text = re.sub(r'http\S+', '', text)
    text = re.sub(r'@[A-Za-z0-9_]+', '', text)
    text = re.sub(r'[^A-Za-zñÑáéíóúÁÉÍÓÚ\s]', ' ', text)

    # Convertir a minúsculas y tokenizar
    words = text.lower().split()
    words = [word for word in words if word not in stopwords.words('spanish')]

    # Lematizar las palabras
    lemmatizer = WordNetLemmatizer()
    words = [lemmatizer.lemmatize(word) for word in words]

    return ' '.join(words)


# Función para obtener las etiquetas de sentimiento del diccionario principal
def get_sentiment_labels(text):
    count_bueno = sum(word in diccionario_principal["palabras_muy_bueno"] for word in text.split())
    count_mal = sum(word in diccionario_principal["palabra_muy_mal"] for word in text.split())

    print("Palabras en el tweet:", text)
    print("Conteo de palabras positivas:", count_bueno)
    print("Conteo de palabras negativas:", count_mal)

    if count_bueno > count_mal:
        return "bueno"
    elif count_mal > count_bueno:
        return "malo"
    else:
        return "neutral"


# Función para leer los tweets de un archivo de Excel
def read_tweets_from_excel(file_path):
    # Cargar el libro de trabajo
    wb = openpyxl.load_workbook(file_path)

    # Obtener la hoja de cálculo activa
    sheet = wb.active

    # Obtener los tweets de la columna 'A'
    tweets = [cell.value for cell in sheet['A']]

    return tweets


# Función para procesar los tweets y obtener los sentimientos
def process_tweets_and_get_sentiments(tweets):
    # Preprocesar los tweets
    preprocessed_tweets = [preprocess_text(str(tweet)) for tweet in tweets]

    # Obtener las etiquetas de sentimiento
    sentiment_labels = [get_sentiment_labels(words) for words in preprocessed_tweets]

    return sentiment_labels


# Función para calcular los porcentajes de sentimientos
def calculate_sentiment_percentages(sentiment_labels1, sentiment_labels2):
    # Calcular los totales de tweets
    total_tweets1 = len(sentiment_labels1)
    total_tweets2 = len(sentiment_labels2)

    # Contar las etiquetas de sentimientos en las matrices numpy
    unique_labels1, counts1 = np.unique(sentiment_labels1, return_counts=True)
    unique_labels2, counts2 = np.unique(sentiment_labels2, return_counts=True)

    # Crear diccionarios para facilitar el acceso a los recuentos de etiquetas
    counts_dict1 = dict(zip(unique_labels1, counts1))
    counts_dict2 = dict(zip(unique_labels2, counts2))

    # Obtener los números de tweets positivos, negativos y neutrales para la primera persona
    positive_tweets1 = counts_dict1.get("bueno", 0)
    negative_tweets1 = counts_dict1.get("malo", 0)
    neutral_tweets1 = total_tweets1 - positive_tweets1 - negative_tweets1

    # Obtener los números de tweets positivos, negativos y neutrales para la segunda persona
    positive_tweets2 = counts_dict2.get("bueno", 0)
    negative_tweets2 = counts_dict2.get("malo", 0)
    neutral_tweets2 = total_tweets2 - positive_tweets2 - negative_tweets2

    # Calcular los porcentajes de sentimientos para la primera persona
    positive_percent1 = (positive_tweets1 / total_tweets1) * 100
    negative_percent1 = (negative_tweets1 / total_tweets1) * 100
    neutral_percent1 = (neutral_tweets1 / total_tweets1) * 100

    # Calcular los porcentajes de sentimientos para la segunda persona
    positive_percent2 = (positive_tweets2 / total_tweets2) * 100
    negative_percent2 = (negative_tweets2 / total_tweets2) * 100
    neutral_percent2 = (neutral_tweets2 / total_tweets2) * 100

    return (
        positive_percent1, negative_percent1, neutral_percent1,
        positive_percent2, negative_percent2, neutral_percent2,
        total_tweets1, total_tweets2,
        positive_tweets1, positive_tweets2,
        negative_tweets1, negative_tweets2,
        neutral_tweets1, neutral_tweets2
    )


# Función para visualizar la comparación de sentimientos
def plot_comparison(positive_percent1, negative_percent1, neutral_percent1, positive_percent2, negative_percent2,
                    neutral_percent2, total_tweets1, total_tweets2, positive_tweets1, positive_tweets2,
                    negative_tweets1, negative_tweets2, neutral_tweets1, neutral_tweets2):
    labels = ['Positivos', 'Negativos', 'Neutrales']

    # Valores para cada persona
    values_persona1 = [positive_percent1, negative_percent1, neutral_percent1]
    values_persona2 = [positive_percent2, negative_percent2, neutral_percent2]

    # Crear el gráfico de barras
    x = range(len(labels))
    plt.figure(figsize=(12, 6))  # Aumentamos el ancho para acomodar la tabla

    plt.bar(x, values_persona1, width=0.3, label='Claudia', color='blue', align='center')
    plt.bar(x, values_persona2, width=0.3, label='Xochitl', color='orange', align='edge')

    # Agregar etiquetas y título
    plt.xticks(x, labels)
    plt.ylabel('Porcentaje')
    plt.title('Comparación de sentimientos')
    plt.legend()

    # Datos para la tabla
    table_data = [
        ['Total Tweets', total_tweets1, total_tweets2],
        ['Positivos', positive_tweets1, positive_tweets2],
        ['Negativos', negative_tweets1, negative_tweets2],
        ['Neutrales', neutral_tweets1, neutral_tweets2]
    ]

    # Crear la tabla y agregarla a la figura
    table = plt.table(cellText=table_data, colLabels=['', 'Claudia', 'Xochitl'], loc='right', colWidths=[0.2, 0.2, 0.2])
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1, 1.5)

    # Ajustar los márgenes para evitar que la tabla se superponga con la gráfica
    plt.subplots_adjust(right=0.7)

    # Mostrar el gráfico
    plt.show()



# Función para entrenar el modelo Naive Bayes
def train_naive_bayes_model(X, y):
    # Crear un vectorizador de conteos para vectorizar las características
    vectorizer = CountVectorizer()

    # Convertir todos los documentos de entrada a cadenas de texto
    X_str = [str(doc) for doc in X]

    # Vectorizar las características de los datos de entrada
    X_vectorized = vectorizer.fit_transform(X_str)

    # Entrenar el modelo Naive Bayes
    model = MultinomialNB()
    model.fit(X_vectorized, y)

    # Guardar el modelo entrenado para su uso posterior
    joblib.dump(model, 'naive_bayes_model.pkl')

    # Guardar el vectorizador para que pueda ser reutilizado
    joblib.dump(vectorizer, 'count_vectorizer.pkl')

    return model, vectorizer

def clean_text(text):
    # Replace accented characters with non-accented equivalents
    text = unidecode(text)
    # Define a pattern to match any character that is not a letter or space
    pattern = re.compile(r'[^A-Za-z ]')
    # Replace the matched characters with a space
    return pattern.sub(' ', text)

def extract_words_from_excel(input_file):
    words = []
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):  # Ensure the cell value is a string
                cleaned_text = clean_text(cell.value)
                words.extend(cleaned_text.lower().split())

    return words

def write_top_words_to_csv(words, output_file):
    # Count the frequency of each word
    word_counts = Counter(words)
    # Select the top 100 most common words
    top_words = [(word, count) for word, count in word_counts.items() if word not in stop_words and len(word) > 3]
    top_words.sort(key=lambda x: x[1], reverse=True)
    top_words = top_words[:100]

    with open(output_file, mode='w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Word', 'Frequency'])
        for word, frequency in top_words:
            writer.writerow([word, frequency])



# Leer los tweets de los archivos de Excel
file_path1 = "C:/Users/Jesus/Documents/Python_Projects/Diccionario2.0/Tiempo1ClaudiaAnalisis.xlsx"
file_path2 = "C:/Users/Jesus/Documents/Python_Projects/Diccionario2.0/Tiempo1XochitlAnalisis.xlsx"

# Uso del programa
input_xlsx = 'Tiempo1ClaudiaAnalisis.xlsx'
input_xlsx2 = 'Tiempo1XochitlAnalisis.xlsx'
output_csv2 = 'conteoPreClaudia.csv'
output_csv = 'conteoPreXochitl.csv'

# Limpia el archivo Excel y extrae las palabras
words = extract_words_from_excel(input_xlsx)
words2 = extract_words_from_excel(input_xlsx2)

# Escribe las 100 palabras más comunes en un archivo CSV
write_top_words_to_csv(words2, output_csv2)
write_top_words_to_csv(words, output_csv)

tweets1 = read_tweets_from_excel(file_path1)
tweets2 = read_tweets_from_excel(file_path2)

# Procesar los tweets y obtener las etiquetas de sentimiento
sentiment_labels1 = process_tweets_and_get_sentiments(tweets1)
sentiment_labels2 = process_tweets_and_get_sentiments(tweets2)

# Calcular los porcentajes de sentimientos
resultados = calculate_sentiment_percentages(sentiment_labels1, sentiment_labels2)

# Extraer los valores de la tupla devuelta
(positive_percent1, negative_percent1, neutral_percent1,
 positive_percent2, negative_percent2, neutral_percent2,
 total_tweets1, total_tweets2,
 positive_tweets1, positive_tweets2,
 negative_tweets1, negative_tweets2,
 neutral_tweets1, neutral_tweets2) = resultados

# Visualizar la comparación de sentimientos
plot_comparison(positive_percent1, negative_percent1, neutral_percent1,
                positive_percent2, negative_percent2, neutral_percent2,
                total_tweets1, total_tweets2,
                positive_tweets1, positive_tweets2,
                negative_tweets1, negative_tweets2,
                neutral_tweets1, neutral_tweets2)

# Obtener los datos de entrada y etiquetas
X = tweets1 + tweets2
y = sentiment_labels1 + sentiment_labels2

# Entrenar el modelo Naive Bayes
train_naive_bayes_model(X, y)

# Leer los tweets de los archivos de Excel para "Tiempo 2"
file_path3 = "C:/Users/Jesus/Documents/Python_Projects/Diccionario2.0/Tiempo2ClaudiaAnalisis.xlsx"
file_path4 = "C:/Users/Jesus/Documents/Python_Projects/Diccionario2.0/Tiempo2XochitlAnalisis.xlsx"

# Uso del programa
input_xlsx = 'Tiempo2ClaudiaAnalisis.xlsx'
input_xlsx2 = 'Tiempo2XochitlAnalisis.xlsx'
output_csv2 = 'conteoPostClaudia.csv'
output_csv = 'conteoPostXochitl.csv'

# Limpia el archivo Excel y extrae las palabras
words = extract_words_from_excel(input_xlsx)
words2 = extract_words_from_excel(input_xlsx2)

# Escribe las 100 palabras más comunes en un archivo CSV
write_top_words_to_csv(words2, output_csv2)
write_top_words_to_csv(words, output_csv)
import joblib

# Cargar el modelo Naive Bayes y el vectorizador guardados
model = joblib.load('naive_bayes_model.pkl')
vectorizer = joblib.load('count_vectorizer.pkl')

# Lee los tweets de los archivos de Excel para "Tiempo 2"
tweets3 = read_tweets_from_excel(file_path3)
tweets4 = read_tweets_from_excel(file_path4)

# Preprocesar los tweets de "Tiempo 2"
preprocessed_tweets3 = [preprocess_text(str(tweet)) for tweet in tweets3]
preprocessed_tweets4 = [preprocess_text(str(tweet)) for tweet in tweets4]

# Vectorizar los tweets de "Tiempo 2" utilizando el vectorizador cargado
X_vectorized3 = vectorizer.transform(preprocessed_tweets3)
X_vectorized4 = vectorizer.transform(preprocessed_tweets4)

# Predecir los sentimientos de los tweets de "Tiempo 2" utilizando el modelo cargado
sentiment_labels3 = model.predict(X_vectorized3)
sentiment_labels4 = model.predict(X_vectorized4)

# Calcular los porcentajes de sentimientos para "Tiempo 2"
resultados_tiempo2 = calculate_sentiment_percentages(sentiment_labels3, sentiment_labels4)

# Extraer los valores de la tupla devuelta para "Tiempo 2"
(positive_percent3, negative_percent3, neutral_percent3,
 positive_percent4, negative_percent4, neutral_percent4,
 total_tweets3, total_tweets4,
 positive_tweets3, positive_tweets4,
 negative_tweets3, negative_tweets4,
 neutral_tweets3, neutral_tweets4) = resultados_tiempo2

# Visualizar la comparación de sentimientos para "Tiempo 2"
plot_comparison(positive_percent3, negative_percent3, neutral_percent3,
                positive_percent4, negative_percent4, neutral_percent4,
                total_tweets3, total_tweets4,
                positive_tweets3, positive_tweets4,
                negative_tweets3, negative_tweets4,
                neutral_tweets3, neutral_tweets4)